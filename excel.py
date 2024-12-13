import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
import openpyxl
import json



def itens_clientes(cliente,hora,cat,dia):
    with open("relatorio_diario_completo.json", "r") as r:
        rel_completo = json.load(r)
    relatorio = rel_completo[dia]

    valores = relatorio[hora][cliente][cat]
    dado = {}

    # NOMES
    cliente_split = str(cliente).split("_")
    nome = ""
    for p in cliente_split:
        nome += " " + p
    dado["Nome"] = nome.title().strip()

    # HORÁRIOS
    dado["Horário"] = hora

    # ITENS
    for e, v in enumerate(valores):
        v_split = str(v).split()
        del v_split[:2]

        v_format = ""
        for p in v_split:
            v_format += " " + p

        dado[f"Item {e + 1}"] = v_format

    valores_add = relatorio[hora][cliente]["ADICIONADO"][cat]
    for e, v in enumerate(valores_add):
        v_split = str(v).split()
        del v_split[:2]

        v_format = ""
        for p in v_split:
            v_format += " " + p

        dado[f"Item {len(dado)-1}"] = v_format

    return dado

def itens_entregas(cliente,hora,endereco,dia):
    with open("relatorio_diario_completo.json", "r") as r:
        rel_completo = json.load(r)
    relatorio = rel_completo[dia]

    categorias = ['salgado', 'doce', 'bolo', 'sobremesa', 'fio_de_ovos']
    dado = {}
    item = 0

    for c in categorias:
        valores = relatorio[hora][cliente][c]

        # NOMES
        cliente_split = str(cliente).split("_")
        nome = ""
        for p in cliente_split:
            nome += " " + p
        dado["Nome"] = nome.title().strip()

        # HORÁRIOS
        dado["Horário"] = hora

        # ENDEREÇO
        dado["Endereço"] = endereco

        # ITENS
        for v in (valores):
            item += 1
            v_split = str(v).split()
            del v_split[:2]

            v_format = ""
            for p in v_split:
                v_format += " " + p

            dado[f"Item {item}"] = v_format

        valores_add = relatorio[hora][cliente]["ADICIONADO"][c]
        for v in (valores_add):
            item += 1
            v_split = str(v).split()
            del v_split[:2]

            v_format = ""
            for p in v_split:
                v_format += " " + p

            dado[f"Item {item}"] = v_format

    return dado

def tabela_encomendas(dia):
    with open("relatorio_diario_completo.json", "r") as r:
        rel_completo = json.load(r)
    relatorio = rel_completo[dia]

    ## ---------------------------- TABELA ----------------------------#

    data_arquivo = dia.replace("/","_")
    nome_arquivo = f"Relatorio_completo_{data_arquivo}.xlsx"

    current_directory = os.getcwd()

    file_path = os.path.join(current_directory,nome_arquivo)

    if os.path.exists(file_path):
        print("Relatório já criado")
        os.startfile(nome_arquivo)
    else:
        categorias = ['doce', 'salgado', 'bolo', 'sobremesa', 'fio_de_ovos']

        with pd.ExcelWriter(nome_arquivo) as writer:
            for categoria in categorias:
                tabela_categoria = pd.DataFrame(columns=["Nome", "Horário"] + [f"Item {i + 1}" for i in range(30)])

                for hora in relatorio:
                    for nome in relatorio[hora]:

                        # Define os dados de cada categoria
                        dados_categoria = itens_clientes(nome, hora, categoria,dia)

                        # Garantir que todos os itens tenham valor
                        for i in range(1, 31):
                            dados_categoria.setdefault(f"Item {i}", "")

                        add = pd.DataFrame([dados_categoria])
                        tabela_categoria = pd.concat([tabela_categoria, add],ignore_index=True)

                tabela_categoria.to_excel(writer, sheet_name=categoria.capitalize(), index=False)

    ## ---------------------------- PERSONALIZAÇÃO ----------------------------#

    wb = openpyxl.load_workbook(nome_arquivo)

    for categoria in wb.sheetnames:
        # try:
        sheet = wb[categoria]

        # Fonte
        fonte = Font(name='Arial', size=12, bold=True)
        # Cor da célula
        fill_header = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # Alinhar
        alignment = Alignment(horizontal="center", vertical="center")

        # Altura das células
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = fonte
                cell.alignment = alignment
                sheet.row_dimensions[(row[0].row)].height = 50

        # Colorir o cabeçalho
        for cell in sheet[1]:
            cell.fill = fill_header

        # Largura das células
        for col in sheet.columns:
            column = col[0].column_letter
            sheet.column_dimensions[column].width = 50

    wb.save(nome_arquivo)

    # Abre o arquivo
    os.startfile(nome_arquivo)

def tabela_entregas(dia):
    with open("relatorio_diario_completo.json", "r") as r:
        rel_completo = json.load(r)
    relatorio = rel_completo[dia]

    ## ---------------------------- TABELA ----------------------------#

    data_arquivo = dia.replace("/", "_")
    nome_arquivo = f"Relatorio_entregas_{data_arquivo}.xlsx"

    current_directory = os.getcwd()

    file_path = os.path.join(current_directory, nome_arquivo)

    if os.path.exists(file_path):
        print("Relatório já criado")
        os.startfile(nome_arquivo)
    else:
        tabela = pd.DataFrame(columns=["Nome", "Horário", "Endereço"] + [f"Item {i + 1}" for i in range(30)])

        for hora in relatorio:
            for cliente in relatorio[hora]:

                endereco = relatorio[hora][cliente]['entrega']

                if endereco != "-":
                    dados = itens_entregas(cliente, hora, endereco, dia)

                    for i in range(1, 31):
                        dados.setdefault(f"Item {i}", "")

                    add = pd.DataFrame([dados])
                    tabela = pd.concat([tabela, add], ignore_index=True)

        with pd.ExcelWriter(nome_arquivo) as w:
            tabela.to_excel(w, sheet_name='Encomendas', index=False)

        ## ---------------------------- PERSONALIZAÇÃO ----------------------------#
        wb = openpyxl.load_workbook(nome_arquivo)
        sheet = wb.active

        # Fonte
        fonte = Font(name='Arial', size=12, bold=True)
        # Cor da célula
        fill_header = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        # Alinhar
        alignment = Alignment(horizontal="center", vertical="center")

        # Altura das células
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = fonte
                cell.alignment = alignment
                sheet.row_dimensions[row[0].row].height = 100

        # Colorir o cabeçalho
        for cell in sheet[1]:
            cell.fill = fill_header

        # Largura das células
        for col in sheet.columns:
            column = col[0].column_letter
            sheet.column_dimensions[column].width = 50

        wb.save(nome_arquivo)
        # Abre o arquivo
        os.startfile(nome_arquivo)


# tabela_encomendas("22/04/2006")
# tabela_entregas("22/04/2006")